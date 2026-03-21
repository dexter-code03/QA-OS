from __future__ import annotations

import re
from pathlib import Path
from typing import Any


def _parse_mobile_call(line: str) -> tuple[str, str] | None:
    """Extract Mobile.<keyword>(<balanced-args>) from a single line."""
    line = line.rstrip(";").strip()
    m = re.match(r"Mobile\.(\w+)\s*\(", line)
    if not m:
        return None
    start_paren = line.find("(", m.start())
    depth = 0
    for i in range(start_paren, len(line)):
        if line[i] == "(":
            depth += 1
        elif line[i] == ")":
            depth -= 1
            if depth == 0:
                return m.group(1), line[start_paren + 1 : i].strip()
    return None


def _selector_from_mobile_args(args: str, aliases: dict[str, str]) -> str | None:
    """Locator short name from inline findTestObject('...') or TestObject variable."""
    m = re.search(r"findTestObject\(\s*['\"]([^'\"]+)['\"]\s*\)", args)
    if m:
        return m.group(1).split("/")[-1]
    m2 = re.match(r"^\s*(\w+)\s*,", args)
    if m2 and m2.group(1) in aliases:
        return aliases[m2.group(1)].split("/")[-1]
    return None


def parse_groovy(source: str) -> list[dict[str, Any]]:
    """
    Parse Katalon Groovy scripts into platform step dicts.
    Handles TestObject aliases, Mobile.* with nested parens, callTestCase placeholders,
    scrollToText, switchToNative, base.takeScreenshot / takeScreeshot typo, retryWait.
    """
    aliases: dict[str, str] = {}
    steps: list[dict] = []

    for raw_line in source.splitlines():
        line = raw_line.strip()
        if not line or line.startswith("//"):
            continue
        if line.startswith("import"):
            continue

        bind = re.match(
            r"^(?:TestObject\s+|def\s+)?(\w+)\s*=\s*findTestObject\(\s*['\"]([^'\"]+)['\"]\s*\)\s*;?\s*$",
            line,
        )
        if bind:
            aliases[bind.group(1)] = bind.group(2)
            continue

        if re.search(r"base\.takeScrees?hot\s*\(", line, re.I):
            steps.append({"type": "takeScreenshot"})
            continue

        retry = re.search(
            r"base\.retryWaitForElementPresent\s*\(\s*findTestObject\(\s*['\"]([^'\"]+)['\"]\s*\)\s*,\s*(\d+)\s*,",
            line,
        )
        if retry:
            name = retry.group(1).split("/")[-1]
            ms = int(retry.group(2)) * 1000
            steps.append({"type": "waitForVisible", "selector": {"using": "accessibilityId", "value": name}, "ms": ms})
            continue

        if re.match(r"base\.\w+\s*\(", line):
            continue

        call = _parse_mobile_call(line)
        if not call:
            continue
        keyword, args_raw = call

        if keyword == "comment":
            continue
        if keyword == "closeApplication":
            continue
        if keyword == "switchToNative":
            steps.append({"type": "wait", "ms": 400})
            continue
        if keyword == "callTestCase":
            steps.append({"type": "wait", "ms": 1500})
            continue
        if keyword == "scrollToText":
            m = re.search(r"['\"]([^'\"]{1,200})['\"]", args_raw)
            label = (m.group(1) if m else "scroll")[:120]
            steps.append({"type": "swipe", "text": "down"})
            steps.append(
                {
                    "type": "waitForVisible",
                    "selector": {"using": "accessibilityId", "value": label},
                    "ms": 8000,
                }
            )
            continue

        handler = {
            "tap": _parse_tap,
            "setText": _parse_set_text,
            "waitForElementPresent": _parse_wait_visible,
            "verifyElementText": _parse_assert_text,
            "verifyElementVisible": _parse_assert_visible,
            "swipe": _parse_swipe,
            "delay": _parse_wait,
            "takeScreenshot": _parse_screenshot,
            "hideKeyboard": _parse_hide_keyboard,
        }.get(keyword)
        if handler:
            step = handler(args_raw, aliases)
            if step:
                steps.append(step)

    return steps


def katalon_or_leaves_and_aliases(source: str) -> tuple[set[str], dict[str, str]]:
    """
    Collect Object Repository leaf names and variable → leaf mappings from a Katalon script.
    Used to snap AI-imported steps back to real locator tokens (not visible UI labels).
    """
    leaves: set[str] = set()
    var_to_leaf: dict[str, str] = {}
    for raw_line in source.splitlines():
        line = raw_line.strip()
        if not line or line.startswith("//"):
            continue
        bind = re.match(
            r"^(?:TestObject\s+|def\s+)?(\w+)\s*=\s*findTestObject\(\s*['\"]([^'\"]+)['\"]\s*\)\s*;?\s*$",
            line,
        )
        if bind:
            leaf = bind.group(2).split("/")[-1]
            var_to_leaf[bind.group(1)] = leaf
            leaves.add(leaf)
            continue
        for m in re.finditer(r"findTestObject\(\s*['\"]([^'\"]+)['\"]\s*\)", line):
            leaves.add(m.group(1).split("/")[-1])
    return leaves, var_to_leaf


def _extract_object_name(args: str, aliases: dict[str, str] | None = None) -> str | None:
    """Pull accessibilityId-style name from findTestObject or aliased variable."""
    if aliases is not None:
        return _selector_from_mobile_args(args, aliases)
    m = re.search(r"findTestObject\(['\"]([^'\"]+)['\"]\)", args)
    if not m:
        return None
    path = m.group(1)
    return path.split("/")[-1]


def _parse_tap(args: str, aliases: dict[str, str]) -> dict | None:
    name = _extract_object_name(args, aliases)
    if not name:
        return None
    return {"type": "tap", "selector": {"using": "accessibilityId", "value": name}}


def _parse_set_text(args: str, aliases: dict[str, str]) -> dict | None:
    name = _extract_object_name(args, aliases)
    if not name:
        return None
    strings = re.findall(r"['\"]([^'\"]*)['\"]", args)
    if len(strings) >= 2:
        text = strings[1]
    elif len(strings) == 1:
        text = strings[0]
    else:
        text = ""
    return {"type": "type", "selector": {"using": "accessibilityId", "value": name}, "text": text}


def _parse_wait_visible(args: str, aliases: dict[str, str]) -> dict | None:
    name = _extract_object_name(args, aliases)
    timeout_match = re.search(r",\s*(\d+)\s*$", args.strip())
    ms = int(timeout_match.group(1)) * 1000 if timeout_match else 10000
    if not name:
        return None
    return {"type": "waitForVisible", "selector": {"using": "accessibilityId", "value": name}, "ms": ms}


def _parse_assert_text(args: str, aliases: dict[str, str]) -> dict | None:
    name = _extract_object_name(args, aliases)
    if not name:
        return None
    strings = re.findall(r"['\"]([^'\"]*)['\"]", args)
    if len(strings) >= 2:
        expected = strings[1]
    elif len(strings) == 1:
        expected = strings[0]
    else:
        expected = ""
    return {"type": "assertText", "selector": {"using": "accessibilityId", "value": name}, "expect": expected}


def _parse_assert_visible(args: str, aliases: dict[str, str]) -> dict | None:
    name = _extract_object_name(args, aliases)
    if not name:
        return None
    return {"type": "assertVisible", "selector": {"using": "accessibilityId", "value": name}}


def _parse_swipe(args: str, _aliases: dict[str, str]) -> dict | None:
    direction_match = re.search(r"['\"]?(up|down|left|right)['\"]?", args, re.IGNORECASE)
    direction = direction_match.group(1).lower() if direction_match else "up"
    return {"type": "swipe", "text": direction}


def _parse_wait(args: str, _aliases: dict[str, str]) -> dict | None:
    num = re.search(r"(\d+)", args)
    ms = int(num.group(1)) * 1000 if num else 1000
    return {"type": "wait", "ms": ms}


def _parse_screenshot(_args: str, _aliases: dict[str, str]) -> dict:
    return {"type": "takeScreenshot"}


def _parse_hide_keyboard(_args: str, _aliases: dict[str, str]) -> dict:
    return {"type": "hideKeyboard"}


def parse_gherkin(source: str) -> list[dict[str, Any]]:
    """
    Very lightweight Gherkin→steps converter.
    """
    steps: list[dict[str, Any]] = []
    for line in source.splitlines():
        line = line.strip()
        for prefix in ("Given ", "When ", "Then ", "And ", "But "):
            if line.startswith(prefix):
                text = line[len(prefix) :]
                steps.append({"type": "gherkin_raw", "text": text})
                break
    return steps


def parse_test_sheet(file_bytes: bytes, filename: str) -> list[dict[str, Any]]:
    """
    Parse an Excel or CSV test sheet into raw test case dicts.
    """
    import io

    rows: list[dict] = []

    if filename.lower().endswith(".csv"):
        import csv

        reader = csv.DictReader(io.StringIO(file_bytes.decode("utf-8", errors="replace")))
        rows = [_normalise_sheet_row(r) for r in reader]
    else:
        try:
            import openpyxl
        except ImportError as e:
            raise ValueError("openpyxl not installed — run: uv sync (or pip install openpyxl)") from e
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        ws = wb.active
        header_row = next(ws.iter_rows(min_row=1, max_row=1))
        headers = [str(c.value or "").strip().lower() for c in header_row]
        for row in ws.iter_rows(min_row=2, values_only=True):
            rows.append(_normalise_sheet_row(dict(zip(headers, row))))

    return [r for r in rows if r.get("name")]


def _normalise_sheet_row(raw: dict) -> dict:
    """Normalise flexible column names to canonical keys."""

    def find(keys: list[str]) -> str:
        for k in keys:
            for rk, rv in raw.items():
                if k in str(rk).lower() and rv is not None and str(rv).strip():
                    return str(rv).strip()
        return ""

    return {
        "name": find(["test name", "test case", "title", "name"]),
        "steps_description": find(["steps", "step description", "action", "description"]),
        "expected": find(["expected", "expected result", "result", "outcome", "acceptance"]),
        "selector_strategy": find(["selector strategy", "strategy", "locator type"]) or "accessibilityId",
        "selector_value": find(["selector value", "locator value", "element", "selector"]),
        "input_value": find(["input", "value", "data", "text"]),
        "priority": find(["priority", "p1", "severity"]),
    }


def _sheet_slug(text: str, max_len: int = 48) -> str:
    s = re.sub(r"[^a-zA-Z0-9]+", "_", (text or "").strip()).strip("_").lower()
    return (s[:max_len] or "element")[:max_len]


def sheet_row_to_minimal_steps(row: dict[str, Any]) -> list[dict[str, Any]]:
    """One-row sheet: optional selector_value → tap (+ type with input_value)."""
    strategy = row.get("selector_strategy") or "accessibilityId"
    value = (row.get("selector_value") or "").strip()
    steps: list[dict[str, Any]] = []
    if value:
        steps.append({"type": "tap", "selector": {"using": strategy, "value": value}})
        if row.get("input_value"):
            steps.append(
                {
                    "type": "type",
                    "selector": {"using": strategy, "value": value},
                    "text": str(row["input_value"]),
                }
            )
    return steps


def sheet_row_heuristic_steps(row: dict[str, Any]) -> list[dict[str, Any]]:
    """Derive steps from free-text Steps + Expected when AI / locator columns are sparse."""
    strategy = row.get("selector_strategy") or "accessibilityId"
    fallback = (row.get("selector_value") or "").strip()
    desc = (row.get("steps_description") or "").strip()
    expected = (row.get("expected") or "").strip()
    input_val = (row.get("input_value") or "").strip()
    steps: list[dict[str, Any]] = []

    if fallback:
        steps.append({"type": "tap", "selector": {"using": strategy, "value": fallback}})
        if input_val:
            steps.append(
                {"type": "type", "selector": {"using": strategy, "value": fallback}, "text": input_val[:500]}
            )

    if not desc:
        if expected:
            sel_final = fallback or _sheet_slug(expected)
            steps.append(
                {"type": "assertText", "selector": {"using": strategy, "value": sel_final}, "expect": expected[:500]}
            )
        return steps

    raw_lines = re.split(r"[\n\r]+|;(?=\s)", desc)
    lines: list[str] = []
    for ln in raw_lines:
        t = re.sub(r"^\s*(\d+[\).\]]|[•\-*◦])\s*", "", ln.strip())
        if t:
            lines.append(t)
    if not lines:
        lines = [desc]

    for line in lines:
        low = line.lower()
        mq = re.search(r"[`'\"]([^`'\"]{1,120})[`'\"]", line)
        quoted = mq.group(1).strip() if mq else None

        if any(
            p in low
            for p in (
                "verify ",
                "assert ",
                "check ",
                "validate ",
                "should show",
                "should display",
                "should see",
                " must ",
                "confirm ",
            )
        ):
            exp = (quoted or expected or line).strip()
            selv = fallback if fallback else _sheet_slug(quoted or line or exp)
            steps.append(
                {"type": "assertText", "selector": {"using": strategy, "value": selv}, "expect": exp[:500]}
            )
            continue

        if any(p in low for p in ("enter ", "type ", "input ", "fill ", "set ")):
            selv = fallback if fallback else _sheet_slug(quoted or line)
            txt = (quoted or input_val or "").strip() or "sample_text"
            steps.append({"type": "tap", "selector": {"using": strategy, "value": selv}})
            steps.append({"type": "type", "selector": {"using": strategy, "value": selv}, "text": str(txt)[:500]})
            continue

        if any(p in low for p in ("tap ", "click ", "press ", "select ", "open ", "go to ", "navigate")):
            selv = fallback if fallback else _sheet_slug(quoted or line)
            steps.append({"type": "tap", "selector": {"using": strategy, "value": selv}})
            continue

        if any(p in low for p in ("wait", "pause", "sleep")):
            msec = 1000
            mn = re.search(r"(\d+)\s*(ms|sec|second|s)\b", low)
            if mn:
                v = int(mn.group(1))
                u = mn.group(2)
                msec = v if u == "ms" else v * 1000
            steps.append({"type": "wait", "ms": min(max(msec, 100), 60_000)})
            continue

        selv = fallback if fallback else _sheet_slug(quoted or line)
        steps.append({"type": "tap", "selector": {"using": strategy, "value": selv}})

    if expected:
        last_assert_expect = None
        for s in reversed(steps):
            if s.get("type") == "assertText":
                last_assert_expect = str(s.get("expect") or "")
                break
        if last_assert_expect != expected:
            sel_final = fallback or _sheet_slug(expected)
            steps.append(
                {"type": "assertText", "selector": {"using": strategy, "value": sel_final}, "expect": expected[:500]}
            )

    return steps


def sheet_row_combined_steps(row: dict[str, Any]) -> list[dict[str, Any]]:
    """Minimal locator-based steps, else heuristic from description + expected."""
    minimal = sheet_row_to_minimal_steps(row)
    if minimal:
        expected = (row.get("expected") or "").strip()
        if expected and not any(s.get("type") in ("assertText", "assertVisible") for s in minimal):
            strategy = row.get("selector_strategy") or "accessibilityId"
            fallback = (row.get("selector_value") or "").strip()
            sel_final = fallback or _sheet_slug(expected)
            minimal = [
                *minimal,
                {"type": "assertText", "selector": {"using": strategy, "value": sel_final}, "expect": expected[:500]},
            ]
        return minimal
    return sheet_row_heuristic_steps(row)


def group_steps_into_test_cases(steps: list[dict[str, Any]], filename: str) -> list[dict[str, Any]]:
    """
    One Groovy / script file → exactly one test case (all steps in order).
    Name = file stem (e.g. Libi 1.groovy → "Libi 1").
    """
    if not steps:
        return []

    stem = Path(filename).stem
    return [
        {
            "name": stem,
            "steps": list(steps),
            "acceptance_criteria": "",
            "import": True,
        }
    ]
